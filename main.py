import os
import pandas as pd
from flask import Flask, request, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def generate_report(input_path, output_path):
    df = pd.read_excel(input_path)

    if not {'Region', 'Client', 'Amount'}.issubset(df.columns):
        raise ValueError("Excel must contain 'Region', 'Client', and 'Amount' columns")

    wb = Workbook()
    wb.remove(wb.active)
    regions = df["Region"].unique()

    for region in regions:
        ws = wb.create_sheet(title=region)
        region_df = df[df["Region"] == region]
        pivot_df = region_df.groupby("Client", as_index=False)["Amount"].sum()
        total = pivot_df["Amount"].sum()

        ws.append(["Client", "Amount"])
        for _, row in pivot_df.iterrows():
            ws.append([row["Client"], row["Amount"]])

        total_row = len(pivot_df) + 2
        ws.cell(row=total_row, column=1, value="Total Sales:")
        ws.cell(row=total_row, column=2, value=total)
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=2).font = Font(bold=True)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="0000CC")

        chart = BarChart()
        chart.title = f"{region} - Sales by Client"
        chart.x_axis.title = "Client"
        chart.y_axis.title = "Amount"
        data = Reference(ws, min_col=2, min_row=1, max_row=1 + len(pivot_df))
        cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(pivot_df))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 10
        ws.add_chart(chart, "E2")

    wb.save(output_path)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files["file"]
        if file and file.filename.endswith(".xlsx"):
            filename = secure_filename(file.filename)
            input_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            output_path = os.path.join(app.config["UPLOAD_FOLDER"], f"report_{filename}")
            file.save(input_path)
            generate_report(input_path, output_path)
            return send_file(output_path, as_attachment=True)
        else:
            return "❌ Please upload a valid .xlsx file."
    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
